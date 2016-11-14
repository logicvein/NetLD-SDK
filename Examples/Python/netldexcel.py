from openpyxl import Workbook, load_workbook
from jsonrpc import JsonRpcProxy, JsonError
from netldutils import NetLdUtils

class NetLdExcel:
   '''A utility class with NetLD/Excel functions'''

   def __init__(self, netldService):
      self._netld_svc = netldService

   def resolve_from_excel(self, excelFile, networks=None, adapters=None):
      r'''A method to resolve devices from an Excel (.xlsx) file.  Resolved
          devices can be restricted to specified networks and/or specified
          adapters.

          The Excel file is assumed to have device IP addresses in the 'A'
          column, with a possible "IP Address" header value in A1 cell.

          The Excel file may also specify the device network in the 'B' column,
          in which case the B1 cell MUST contain the text "Network".

          :param excelFile: the file name of the Excel file to resolve from.
          :type excelFile: string

          :param networks: a single network name, list of names, or None
          :type networks: string or list
      '''

      workbook = load_workbook(excelFile)
      ws = workbook.active

      cols = self.__detect_excel_columns(workbook)
      networkColumn = cols['networkColumn']
      startRow = cols['startRow']

      devices = {}
      netDict = {}
      network = None
      for row in range(startRow, ws.max_row + 1):
         ipAddress = ws.cell(row=row, column=1).value
         key = ipAddress
         if networkColumn:
            network = ws.cell(row=row, column=networkColumn).value
            netDict[network] = True
            key = ipAddress + network
         devices[key] = {'ipAddress': ipAddress, 'network': network}

      if networks is None and networkColumn:
         networks = netDict.keys()
      elif networks is None:
         networks = NetLdUtils(netld_svc).get_networks()
      elif not isinstance(networks, list):
         networks = [networks]

      pageData= {'offset': 0, 'pageSize': 1000}
      while True:
         pageData = self._netld_svc.call('Inventory.search', networks, 'ipAddress', '', pageData, 'ipAddress', False)
         for invDevice in pageData['devices']:
            ipAddress = invDevice['ipAddress']
            network = invDevice['network']
            adapter = invDevice['adapterId']
            if adapters and not adapter in adapters:
               continue

            key = ipAddress + network
            device = devices.get(key)
            if device is None:
               device = devices.get(ipAddress)
               if device is None:
                  continue
            if device['network'] == network or not 'resolved' in device:
               invDevice['resolved'] = True
               devices[key] = invDevice

         if pageData['offset'] + pageData['pageSize'] > pageData['total']:
            break;
         else:
            pageData['offset'] += pageData['pageSize']

      for key, device in devices.items():
         if not 'resolved' in device:
            print 'Unresolved device ' + device['ipAddress']
            del devices[key]

      return devices.values()

   def __detect_excel_columns(self, workbook):
      startRow = 1
      netColumn = None
      ws = workbook.active
      if (ws['A1'].value == 'IP Address'):
         startRow = 2
      else:
         try:
            ipaddress.ip_address(ws['A1'].value)
         except ValueError:
            raise ValueError('IP address column could not be identified.')

      for row in ws.iter_rows(min_row=1, max_row=1):
         for cell in row:
            if (cell.value == 'Network'):
               netColumn = cell.col_idx

      return {'networkColumn': netColumn, 'startRow': startRow}
