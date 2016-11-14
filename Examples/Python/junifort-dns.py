#!/usr/bin/env python

# Non-default required modules: ipaddress, openpyxl
import os
import re
import sys
import time
import getopt
import ipaddress
import ConfigParser
from openpyxl import Workbook, load_workbook
from jsonrpc import JsonRpcProxy, JsonError
from netldutils import NetLdUtils
from netldexcel import NetLdExcel


juniper_read_template = """
get nsm | i server
get conf | i "host dns[1-3]"
"""

juniper_nsm_regex = re.compile(r"(?:host: (\S+),)")
juniper_dns_regex = re.compile(r"(?:dns[1-3] (\S+))")

juniper_write_template = """
set dns host dns1 {DNS-X1}
set dns host dns2 {DNS-X2}
set dns host dns3 {DNS-X3}
set dns server-select domain * primary-server {DNS-X1} secondary-server {DNS-X2} tertiary-server {DNS-X3} failover
save
"""

juniper_validation_regex = re.compile(r"\S+->.*?[\r\n](?=(\S+?->|\s+))")

fortinet_read_template = """
show | grep 'set fmg'
config global
config system dns
get
"""

fortinet_fmg_regex = re.compile(r"(?:fmg \"(\S+)\")")
fortinet_dns_regex = re.compile(r"(?:^(?:primary|secondary).*?: (\S+))", re.MULTILINE)

fortinet_write_template = """
config global
config system dns
set primary {DNS-X1}
set secondary {DNS-X2}
end
end
exit
"""

fortinet_validation_regex = re.compile(r"(\S+).*?#.*?[\r\n](?=[\r\n](\1.*?#|\S+))")

def usage_and_exit(argv):
   print 'Usage: ' + sys.argv[0] + ' [OPTIONS]'
   print 'Execute a pre-existing job in Net LineDancer.'
   print
   print '  -f --excel=filename  The name of the Excel file to read/write'
   print '  -e --export Read the Excel file and export the DNS/Mgr settings'
   print '  -x --exec Read an Excel file and apply the DNS/Mgr settings'
   print '  -h host --host=host  Hostname or IP address of Net LineDancer server'
   print '  -u username --user=username  Net LineDancer username'
   print '  -p password --password=password  Net LineDancer password'
   print
   sys.exit(2)

netld_svc = None
netld_user = 'admin'
netld_pass = 'password'
netld_host = 'localhost'

def main(argv):
   try:
      opts, args = getopt.getopt(argv, "h:u:p:f:e:x:", ["host=","user=","password=","excel=","export","exec"])
   except getopt.GetoptError as err:
      print "Error: " + str(err)
      usage_and_exit(argv)

   global netld_host, netld_user, netld_pass, excel_file
   excel_file = None
   export_action = False
   exec_action = False

   if os.path.exists('junifort.ini'):
      config = ConfigParser.RawConfigParser()
      config.read('junifort.ini')
      for section in config.sections():
         netld_host = config.get(section, 'host')
         netld_user = config.get(section, 'username')
         netld_pass = config.get(section, 'password')

   for opt, arg in opts:
      if opt in ('-h', '--host'):
         netld_host = arg
      if opt in ('-u', '--user'):
         netld_user = arg
      if opt in ('-p', '--password'):
         netld_pass = arg
      if opt in ('-f', '--excel'):
         excel_file = arg
      if opt in ('-e', '--export'):
         export_action = True
      if opt in ('-x', '--exec'):
         exec_action = True

   if excel_file is None or (not export_action and not exec_action):
      usage_and_exit(argv)

   try:
      global netld_svc
      netld_svc = JsonRpcProxy.fromHost(netld_host, netld_user, netld_pass)

      devices = NetLdExcel(netld_svc).resolve_from_excel(
         excelFile=excel_file,
         adapters=['Juniper::ScreenOS', 'Fortinet::FortiGate']
      )

      if export_action:
         export_to_excel(devices)
      else:
         exec_from_excel(devices, excel_file)

   except JsonError as ex:
      print 'JsonError: ' + str(ex.value)
   finally:
       netld_svc.call('Security.logoutCurrentUser')

def export_to_excel(devices):
   exportWb = Workbook()
   exportWorksheet = exportWb.active
   exportWorksheet.append(['IP Address', 'Network', 'MGR', 'DNS1', 'DNS2', 'DNS3'])

   utils = NetLdUtils(netld_svc)

   executions = []
   for adapterId in ['Juniper::ScreenOS', 'Fortinet::FortiGate']:
      # Filter devices by adapterId
      job_devices = list(device for device in devices if device['adapterId'] == adapterId)

      if len(job_devices) == 0:
         continue

      try:
         def create_closure(mgr_regex, dns_regex):
            def append_excel_row(ipAddress, network, response):
               result = mgr_regex.search(response)
               if result is None:
                  print "Unexpected response from {0} in network '{1}'".format(ipAddress, network)
               mgr = result.group(1) if result else ''
               dns = dns_regex.findall(response)

               exportWorksheet.append([
                  ipAddress,
                  network,
                  mgr if mgr else '',
                  dns[0] if len(dns) > 0 else '',
                  dns[1] if len(dns) > 1 else '',
                  dns[2] if len(dns) > 2 else '']
               )
            # return the closure function
            return append_excel_row

         executions.append({
            'execution': utils.execute_command_runner(
               devices  = job_devices,
               name     = 'Export ' + adapterId + ' DNS/Manager to Excel',
               commands = juniper_read_template if adapterId == 'Juniper::ScreenOS' else fortinet_read_template),
            'callback' : create_closure(
               mgr_regex = juniper_nsm_regex if adapterId == 'Juniper::ScreenOS' else fortinet_fmg_regex,
               dns_regex = juniper_dns_regex if adapterId == 'Juniper::ScreenOS' else fortinet_dns_regex)
         })
      except JsonError as ex:
         print 'JsonError: ' + str(ex.value)

   try:
      for execs in executions:
         utils.wait_job_completion(execs['execution'])

         utils.get_tool_details(execs['execution'], execs['callback'])
   except JsonError as ex:
      print 'JsonError: ' + str(ex.value)

   exportWb.save('export.xlsx')


def exec_from_excel(devices, excelFile):
   workbook = load_workbook(excelFile)
   worksheet = workbook.active

   utils = NetLdUtils(netld_svc)

   executions = []
   for adapterId in ['Juniper::ScreenOS', 'Fortinet::FortiGate']:
      # Create a dictionary of Devices objects filtered by adapterId
      job_devices = {(device['ipAddress'] + device['network']): device for device in devices if device['adapterId'] == adapterId}

      if len(job_devices) == 0:
         continue

      # Unique list of networks from the dictionary of job_devices
      networks = { device['network']: True for device in job_devices.values() }.keys()

      replacements = []
      for row_num in range(2, worksheet.max_row + 1):
         ipAddress = worksheet.cell( row=row_num, column=1).value
         network   = worksheet.cell( row=row_num, column=2).value
         unused    = worksheet.cell( row=row_num, column=3).value
         dns1      = worksheet.cell( row=row_num, column=4).value
         dns2      = worksheet.cell( row=row_num, column=5).value
         dns3      = worksheet.cell( row=row_num, column=6).value

         if job_devices.get( ipAddress + network ):
            replacements.append({
               'ipAddress': ipAddress,
               'network': network,
               'DNS-X1': dns1 if dns1 else '',
               'DNS-X2': dns2 if dns2 else dns1,
               'DNS-X3': dns3 if dns3 else dns1,
            })

      job_data = utils.create_smart_change(
         name     = adapterId + ' apply DNS settings',
         template = juniper_write_template if adapterId == 'Juniper::ScreenOS' else fortinet_write_template,
         networks = networks
      )

      try:
         def create_closure(adapter, validation_regex, match_group, match_string):
            def check_execution(ipAddress, network, response):
               for match in validation_regex.finditer(response):
                  if match.group(match_group) and not match_string in match.group(match_group):
                     print "{0} device {1} in network '{2}' did not contain expected response.".format(adapter, ipAddress, network)
            # return the closure function
            return check_execution

         executions.append({
            'execution': utils.execute_smart_change(job_data, job_devices.values(), 'perdevice', replacements),
            'callback' : create_closure(
               adapter = adapterId,
               validation_regex = juniper_validation_regex if adapterId == 'Juniper::ScreenOS' else fortinet_validation_regex,
               match_group  = 1 if adapterId == 'Juniper::ScreenOS' else 2,
               match_string = '->' if adapterId == 'Juniper::ScreenOS' else '#'),
         })
      except JsonError as ex:
         print 'JsonError: ' + str(ex.value)

   try:
      for execs in executions:
         utils.wait_job_completion(execs['execution'])

         utils.get_tool_details(execs['execution'], execs['callback'])

   except JsonError as ex:
      print 'JsonError: ' + str(ex.value)


if __name__ == "__main__":
   exit(main(sys.argv[1:]))
